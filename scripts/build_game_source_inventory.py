#!/usr/bin/env python3
"""Build a canonical, deduped Game News source inventory from a workbook.

This is an intake/research tool, not part of the live fetch pipeline.
It reads a maintainer-owned spreadsheet of source leads, collapses obvious
duplicates, and writes review artifacts under docs/research/ so new source
admission can happen deliberately instead of bulk-importing a noisy list into
the crawler.

The output model is intentionally two-layered:
1. canonical_sources: deduped by exact normalized URL
2. host_families: host-level source families for manual consolidation review

Why both?
- Exact URL dedupe is safe to automate.
- Same-host consolidation often needs judgment. Example:
  `example.com/` and `example.com/news` are probably related, but should not be
  auto-merged blindly if one is a homepage and the other is the real newsroom.
"""
from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

REPO_ROOT = Path(__file__).resolve().parent.parent

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to read the workbook. Install it in your local "
        "dev environment before running this script."
    ) from exc

from game_sources import DIRECT_RSS_SOURCES  # noqa: E402

DEFAULT_WORKBOOK = Path.home() / "Desktop" / "MASTER Gaming Media List (Jun 2026).xlsx"
DEFAULT_JSON = REPO_ROOT / "docs" / "research" / "game-source-canonical-inventory-2026-07-07.json"
DEFAULT_MD = REPO_ROOT / "docs" / "research" / "game-source-canonical-inventory-2026-07-07.md"

SOCIAL_HOST_MARKERS = {
    "facebook.com",
    "instagram.com",
    "twitter.com",
    "x.com",
    "youtube.com",
    "youtu.be",
    "tiktok.com",
    "twitch.tv",
    "reddit.com",
    "discord.gg",
    "discord.com",
    "linkedin.com",
}

COMMUNITY_HOST_MARKERS = {
    "bilibili.com",
    "douyu.com",
    "huya.com",
    "arca.live",
    "ruliweb.com",
    "inven.co.kr",
    "fmkorea.com",
    "op.gg",
}

NEWS_PATH_HINTS = (
    "/news",
    "/news/",
    "/newswire",
    "/press",
    "/blog",
    "/media",
    "/release",
    "/releases",
    "/editorial",
    "/discover",
)

OFFICIAL_NAME_HINTS = (
    "playstation",
    "nintendo",
    "xbox",
    "garena",
    "rockstar",
    "capcom",
    "nexon",
    "ncsoft",
    "krafton",
    "neowiz",
    "com2us",
    "devsisters",
    "pearl abyss",
    "shift up",
    "perfect world",
    "warner bros",
    "square enix",
    "playrix",
    "2k",
    "larian",
    "valve",
    "kakao games",
    "kuro games",
    "gravity",
    "century games",
)

GAME_MEDIA_NAME_HINTS = (
    "game",
    "gaming",
    "gamer",
    "esports",
    "ggwp",
    "geek culture",
    "ungeek",
    "back2gaming",
    "kotakgame",
    "dunia games",
    "gamek",
    "xemgame",
    "thisisgame",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-json", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--output-md", type=Path, default=DEFAULT_MD)
    return parser.parse_args()


def canonicalize_url(raw_url: str) -> dict[str, str]:
    src = (raw_url or "").strip()
    if not src:
        return {"raw": "", "url": "", "host": "", "path": ""}
    if not re.match(r"^https?://", src, re.I):
        src = f"https://{src}"
    parsed = urlparse(src)
    host = (parsed.netloc or "").lower()
    if host.startswith("www."):
        host = host[4:]
    path = (parsed.path or "/").rstrip("/") or "/"
    path = path.lower()
    url = f"{host}{'' if path == '/' else path}"
    return {"raw": raw_url, "url": url, "host": host, "path": path}


def normalize_source_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip())


def slugify(text: str) -> str:
    text = re.sub(r"[^a-zA-Z0-9]+", "-", text.lower()).strip("-")
    return text or "source"


def traffic_int(value: str) -> int | None:
    s = str(value or "").strip().replace(",", "")
    if not s or s == "0":
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def guess_source_class(entry: dict[str, Any]) -> str:
    host = entry["host"]
    name = entry["display_name"].lower()
    if host in SOCIAL_HOST_MARKERS:
        return "social"
    if host in COMMUNITY_HOST_MARKERS:
        return "community"
    if any(hint in name for hint in OFFICIAL_NAME_HINTS):
        return "official_publisher_or_platform"
    if any(hint in name for hint in GAME_MEDIA_NAME_HINTS):
        return "regional_game_media"
    if any(hint in entry["path"] for hint in NEWS_PATH_HINTS):
        return "editorial_or_newsroom"
    return "website_candidate"


def research_action(entry: dict[str, Any], live_hosts: set[str]) -> str:
    if entry["host"] in live_hosts:
        return "already_live"
    if entry["source_class"] == "social":
        return "optional_social_lane"
    if entry["source_class"] == "community":
        return "manual_review_not_default"
    if entry["path"] != "/" and any(hint in entry["path"] for hint in NEWS_PATH_HINTS):
        return "candidate_stable_newsroom"
    return "candidate_feed_discovery"


def crawl_priority(entry: dict[str, Any]) -> str:
    action = entry["research_action"]
    if action == "already_live":
        return "live"
    if action == "candidate_stable_newsroom":
        return "high"
    if action == "candidate_feed_discovery":
        return "medium"
    if action == "optional_social_lane":
        return "optional"
    return "review"


def build_inventory(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        iterator = sheet.iter_rows(values_only=True)
        header = next(iterator, None)
        if not header:
            continue
        for row in iterator:
            if not row or not any(row):
                continue
            cells = list(row) + [None] * 5
            name, url_type, url, traffic, ads = cells[:5]
            parsed = canonicalize_url(str(url or ""))
            rows.append(
                {
                    "sheet_region": str(sheet.title or "").strip(),
                    "name": normalize_source_name(str(name or "")),
                    "url_type": normalize_source_name(str(url_type or "")),
                    "raw_url": str(url or "").strip(),
                    "canonical_url": parsed["url"],
                    "host": parsed["host"],
                    "path": parsed["path"],
                    "traffic_estimate": traffic_int(str(traffic or "")),
                    "ads": normalize_source_name(str(ads or "")),
                }
            )

    exact_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = row["canonical_url"] or f"missing::{slugify(row['name'])}"
        exact_groups[key].append(row)

    live_hosts = {
        canonicalize_url(str(source.get("feed_url") or "")).get("host", "")
        for source in DIRECT_RSS_SOURCES
    }
    live_hosts.discard("")

    canonical_sources: list[dict[str, Any]] = []
    host_family_map: dict[str, list[str]] = defaultdict(list)

    for canonical_url, group in sorted(exact_groups.items()):
        names = [item["name"] for item in group if item["name"]]
        name_counts = Counter(names)
        display_name = name_counts.most_common(1)[0][0] if name_counts else canonical_url
        hosts = [item["host"] for item in group if item["host"]]
        host = Counter(hosts).most_common(1)[0][0] if hosts else ""
        paths = [item["path"] for item in group if item["path"]]
        path = Counter(paths).most_common(1)[0][0] if paths else "/"
        url_type_counts = Counter(item["url_type"] for item in group if item["url_type"])
        url_type = url_type_counts.most_common(1)[0][0] if url_type_counts else "Unknown"
        regions = sorted({item["sheet_region"] for item in group if item["sheet_region"]})
        traffics = [item["traffic_estimate"] for item in group if item["traffic_estimate"] is not None]
        entry = {
            "canonical_source_id": slugify(f"{display_name}-{canonical_url or host}"),
            "display_name": display_name,
            "canonical_url": canonical_url,
            "host": host,
            "path": path,
            "url_type": url_type,
            "sheet_regions": regions,
            "sheet_row_count": len(group),
            "names_seen": sorted({item["name"] for item in group if item["name"]}),
            "raw_urls_seen": sorted({item["raw_url"] for item in group if item["raw_url"]}),
            "traffic_max": max(traffics) if traffics else None,
            "ads_seen": sorted({item["ads"] for item in group if item["ads"]}),
        }
        entry["source_class"] = guess_source_class(entry)
        entry["research_action"] = research_action(entry, live_hosts)
        entry["crawl_priority"] = crawl_priority(entry)
        entry["matches_current_live_source"] = host in live_hosts
        canonical_sources.append(entry)
        host_family_map[host].append(entry["canonical_source_id"])

    family_counts = Counter(row["host"] for row in rows if row["host"])
    family_url_counts = Counter(entry["host"] for entry in canonical_sources if entry["host"])
    host_families: list[dict[str, Any]] = []
    for host, entry_ids in sorted(host_family_map.items(), key=lambda kv: (-len(kv[1]), kv[0])):
        member_entries = [entry for entry in canonical_sources if entry["canonical_source_id"] in entry_ids]
        host_families.append(
            {
                "host": host,
                "canonical_entry_count": len(entry_ids),
                "workbook_row_count": family_counts.get(host, 0),
                "matches_current_live_source": host in live_hosts,
                "source_classes": sorted({entry["source_class"] for entry in member_entries}),
                "research_actions": sorted({entry["research_action"] for entry in member_entries}),
                "sheet_regions": sorted({region for entry in member_entries for region in entry["sheet_regions"]}),
                "display_names": [entry["display_name"] for entry in member_entries[:10]],
                "canonical_source_ids": entry_ids,
            }
        )

    summary = {
        "rows_total": len(rows),
        "website_rows": sum(1 for row in rows if row["url_type"].lower() == "website"),
        "social_rows": sum(1 for row in rows if row["url_type"].lower() == "social media"),
        "non_http_rows": sum(1 for row in rows if row["raw_url"] and not re.match(r"^https?://", row["raw_url"], re.I)),
        "canonical_sources_exact_url": len(canonical_sources),
        "host_families": len(host_families),
        "host_families_multi_entry": sum(1 for family in host_families if family["canonical_entry_count"] > 1),
        "already_live_matches": sum(1 for entry in canonical_sources if entry["matches_current_live_source"]),
        "candidate_stable_newsroom": sum(1 for entry in canonical_sources if entry["research_action"] == "candidate_stable_newsroom"),
        "candidate_feed_discovery": sum(1 for entry in canonical_sources if entry["research_action"] == "candidate_feed_discovery"),
        "optional_social_lane": sum(1 for entry in canonical_sources if entry["research_action"] == "optional_social_lane"),
        "manual_review_not_default": sum(1 for entry in canonical_sources if entry["research_action"] == "manual_review_not_default"),
    }

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "generated_note": (
            "Canonical Game News source inventory built from workbook leads. "
            "Exact URL duplicates are collapsed automatically; same-host families "
            "remain visible for manual consolidation review before any crawler changes."
        ),
        "input_workbook": str(workbook_path),
        "summary": summary,
        "current_live_sources": [
            {
                "site_id": source["site_id"],
                "site_name": source["site_name"],
                "feed_url": source["feed_url"],
                "region": source["region"],
            }
            for source in DIRECT_RSS_SOURCES
        ],
        "canonical_sources": sorted(
            canonical_sources,
            key=lambda item: (
                {"live": 0, "high": 1, "medium": 2, "review": 3, "optional": 4}.get(item["crawl_priority"], 5),
                item["display_name"].lower(),
            ),
        ),
        "host_families": host_families,
    }


def write_markdown(payload: dict[str, Any], path: Path) -> None:
    summary = payload["summary"]
    live_sources = payload["current_live_sources"]
    canonical_sources = payload["canonical_sources"]
    host_families = payload["host_families"]

    top_families = [family for family in host_families if family["canonical_entry_count"] > 1][:20]
    add_now = [item for item in canonical_sources if item["research_action"] == "candidate_stable_newsroom"][:25]
    feed_discovery = [item for item in canonical_sources if item["research_action"] == "candidate_feed_discovery"][:25]
    social = [item for item in canonical_sources if item["research_action"] == "optional_social_lane"][:20]

    lines = [
        "# Game Source Canonical Inventory (2026-07-07)",
        "",
        f"Generated from workbook: `{payload['input_workbook']}`",
        "",
        "## Summary",
        "",
        f"- Workbook rows: `{summary['rows_total']}`",
        f"- Website rows: `{summary['website_rows']}`",
        f"- Social rows: `{summary['social_rows']}`",
        f"- Exact-URL canonical sources: `{summary['canonical_sources_exact_url']}`",
        f"- Host families: `{summary['host_families']}`",
        f"- Host families with multiple canonical entries: `{summary['host_families_multi_entry']}`",
        f"- Already-live matches: `{summary['already_live_matches']}`",
        f"- Stable-newsroom candidates: `{summary['candidate_stable_newsroom']}`",
        f"- Feed-discovery candidates: `{summary['candidate_feed_discovery']}`",
        f"- Optional social-lane entries: `{summary['optional_social_lane']}`",
        "",
        "## Current Live Sources",
        "",
    ]

    for source in live_sources:
        lines.append(f"- `{source['site_name']}` -> `{source['feed_url']}` ({source['region']})")

    lines.extend(
        [
            "",
            "## Highest-Priority Stable Newsroom Candidates",
            "",
            "| Name | Regions | URL | Class | Action |",
            "| --- | --- | --- | --- | --- |",
        ]
    )
    for item in add_now:
        lines.append(
            f"| {item['display_name']} | {', '.join(item['sheet_regions'])} | `{item['canonical_url']}` | "
            f"{item['source_class']} | {item['research_action']} |"
        )

    lines.extend(
        [
            "",
            "## Feed Discovery Candidates",
            "",
            "| Name | Regions | URL | Class | Action |",
            "| --- | --- | --- | --- | --- |",
        ]
    )
    for item in feed_discovery:
        lines.append(
            f"| {item['display_name']} | {', '.join(item['sheet_regions'])} | `{item['canonical_url']}` | "
            f"{item['source_class']} | {item['research_action']} |"
        )

    lines.extend(
        [
            "",
            "## Optional Social Lane",
            "",
            "| Name | Regions | URL | Action |",
            "| --- | --- | --- | --- |",
        ]
    )
    for item in social:
        lines.append(
            f"| {item['display_name']} | {', '.join(item['sheet_regions'])} | `{item['canonical_url']}` | "
            f"{item['research_action']} |"
        )

    lines.extend(
        [
            "",
            "## Host Families Needing Manual Consolidation Review",
            "",
            "| Host | Canonical entries | Workbook rows | Regions | Live? |",
            "| --- | --- | --- | --- | --- |",
        ]
    )
    for family in top_families:
        lines.append(
            f"| `{family['host']}` | {family['canonical_entry_count']} | {family['workbook_row_count']} | "
            f"{', '.join(family['sheet_regions'])} | {'yes' if family['matches_current_live_source'] else 'no'} |"
        )

    lines.extend(
        [
            "",
            "## Intake Rule",
            "",
            "- Do not bulk-import this file into the live crawler.",
            "- Review `candidate_stable_newsroom` first.",
            "- Keep `optional_social_lane` out of the default crawl path.",
            "- Use host-family review to collapse repeated publisher/newsroom families before adding sources.",
            "",
        ]
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    args = parse_args()
    payload = build_inventory(args.workbook)
    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_markdown(payload, args.output_md)
    print(f"Wrote {args.output_json}")
    print(f"Wrote {args.output_md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
